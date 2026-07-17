import streamlit as st
import pandas as pd
import io
from openpyxl.styles import PatternFill, Font

# ---------------- PAGE CONFIG ----------------
st.set_page_config(
    page_title="Bank Reconciliation Tool - Muller & Phipps",
    page_icon="🏦",
    layout="wide"
)

# ---------------- BANK CONFIGURATION ----------------
# To add a new bank later: add an entry here with its sheet name, header
# row (0-indexed), and the column names for deposit slip / amount / depot.
BANKS = {
    "MCB": {
        "account_code": 212201, "account_label": "MCB Bank Ltd",
        "sheet": "MCB", "header": 1,
        "slip_col": "Deposit Slip No.", "amount_col": "Amount",
        "depot_col": "Dealer Code", "date_col": "Credit Date",
    },
    "HBL": {
        "account_code": 212202, "account_label": "Habib Bank Limited",
        "sheet": "HBL", "header": 6,
        "slip_col": "Deposit Slip No.", "amount_col": "Amount",
        "depot_col": "Dealer Code", "date_col": "Credit Date",
    },
    "UBL": {
        "account_code": 212205, "account_label": "United Bank Limited",
        "sheet": "UBL", "header": 0,
        "slip_col": "DEPOSIT SLIP NO", "amount_col": "AMOUNT",
        "depot_col": "DEPOT'S CODE", "date_col": "REALIZATION DATE",
    },
    "FAYSAL": {
        "account_code": 212209, "account_label": "Faysal Bank Limited",
        "sheet": "MIS", "header": 4,
        "slip_col": "Deposit No.", "amount_col": "Amount",
        "depot_col": "Dealer Code", "date_col": "Credit Date",
    },
}

DEPOT_CODES = [612, 613, 614, 651, 657, 645, 646, 733, 766, 778]
TOLERANCE = 1.0  # rupees - absorbs rounding differences


# ---------------- CSS ----------------
st.markdown("""
<style>
div.stDownloadButton > button, div.stButton > button {
    width: 100%;
    background-color: #1f4e79;
    color: white;
    font-size: 16px;
    font-weight: bold;
    padding: 12px;
    border-radius: 10px;
    border: none;
}
div.stDownloadButton > button:hover, div.stButton > button:hover {
    background-color: #163a5c;
}
</style>
""", unsafe_allow_html=True)


# ---------------- HELPERS ----------------

def load_gl(file):
    df = pd.read_excel(file, sheet_name="GL")
    df.columns = [str(c).strip() for c in df.columns]
    return df


def load_bank(file, bank_key):
    cfg = BANKS[bank_key]
    df = pd.read_excel(file, sheet_name=cfg["sheet"], header=cfg["header"])
    df.columns = [str(c).strip() for c in df.columns]
    return df


def reconcile_account(gl_df, bank_df, bank_key, depot_code):
    cfg = BANKS[bank_key]

    gl_sub = gl_df[gl_df["Natural Account Segment"] == cfg["account_code"]].copy()
    dr = gl_sub[gl_sub["Entered Amount DR"].notna()].copy()
    cr = gl_sub[gl_sub["Entered Amount CR"].notna()].copy()

    bank_sub = bank_df[bank_df[cfg["depot_col"]] == depot_code].copy()
    bank_sub["_used"] = False

    dr["Status"] = "UNMATCHED"
    dr["Bank Deposit Slip"] = pd.Series([""] * len(dr), index=dr.index, dtype="object")
    dr["Bank Amount"] = None
    dr["Bank Date"] = pd.Series([""] * len(dr), index=dr.index, dtype="object")

    for idx, row in dr.iterrows():
        amt = row["Entered Amount DR"]
        if pd.isna(amt):
            continue
        candidates = bank_sub[
            (~bank_sub["_used"]) & ((bank_sub[cfg["amount_col"]] - amt).abs() <= TOLERANCE)
        ]
        if len(candidates) > 0:
            match_idx = candidates.index[0]
            bank_sub.loc[match_idx, "_used"] = True
            dr.loc[idx, "Status"] = "MATCHED"
            dr.loc[idx, "Bank Deposit Slip"] = str(bank_sub.loc[match_idx, cfg["slip_col"]])
            dr.loc[idx, "Bank Amount"] = bank_sub.loc[match_idx, cfg["amount_col"]]
            dr.loc[idx, "Bank Date"] = str(bank_sub.loc[match_idx, cfg["date_col"]])

    unmatched_bank = bank_sub[~bank_sub["_used"]].drop(columns=["_used"])

    return {
        "bank": bank_key,
        "account_code": cfg["account_code"],
        "account_label": cfg["account_label"],
        "ledger_debits": dr,
        "ledger_credits": cr,
        "unmatched_bank_items": unmatched_bank,
    }


# Full list of GLR accounts in the exact order used in the real template.
# bank_key links to BANKS dict where we have a parser; None = no data source yet.
FULL_ACCOUNTS = [
    ("MCB Bank Ltd", 212201, "MCB"),
    ("Habib Bank Limited", 212202, "HBL"),
    ("National Bank of Pakistan", 212203, None),
    ("HBL Konnect", 212204, None),
    ("United Bank Limited", 212205, "UBL"),
    ("Bank Al Falah Limited", 212206, None),
    ("Bank Al Habib Limited", 212207, None),
    ("Samba Bank Limited", 212208, None),
    ("Faysal Bank Limited", 212209, "FAYSAL"),
    ("Dubai Islamic Bank Limited", 212210, None),
    ("Meezan Bank Limited", 212211, None),
    ("Habib Metropoliton Bank", 212212, None),
    ("Meezan Bank Limited (MFS)", 212213, None),
    ("Habib Bank Limited (MFS)", 212214, None),
    ("Askari Bank Limited", 212215, None),
    ("Bank of Punjab", 212216, None),
    ("United Bank Limited (MFS)", 212217, None),
    ("Bank Islami Limited", 212218, None),
    ("UBL Omni", 212219, None),
    ("Standard Chartered Bank", 212503, None),
]

ANNEX_LEFT_HEADERS = [
    "S. No.", "Dates", "D/Slip No\nas Per Salesman", "D/Slip No\nas Per\nBank MIS",
    "Ch/Cash", "Chq. No.", "Amount", "Clearance Date", "Remarks"
]


def build_template_excel(results_by_bank, depot_code, depot_name=""):
    """Builds output in the same shape as the real Main Reconciliation /
    Annex 1 (Debit M&P) / Annex 3 (Debit Bank) template."""

    import openpyxl
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    navy_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    navy_font = Font(color="FFFFFF", bold=True)
    red_fill = PatternFill(start_color="FDEAEA", end_color="FDEAEA", fill_type="solid")
    bold = Font(bold=True)

    # results_by_bank: {bank_key: result_dict} only for banks actually uploaded

    # ---------------- MAIN RECONCILIATION ----------------
    ws = wb.active
    ws.title = "Main Reconciliation"

    ws["B2"] = "Muller & Phipps Pakistan (Pvt) Ltd"
    ws["B3"] = "Reconciliation of CMD / Collection Accounts"
    ws["B5"] = "Depot Name:"
    ws["C5"] = depot_code
    ws["E5"] = depot_name

    headers = [
        "Branch Name\nAs Per GLR", "Location\nCode", "GLR\nCode No.",
        "Collection Bank Balance", "Add Debit in M&P Ledger\n(Annexure 1)",
        "Less Credit in M&P Ledger\n(Annexure 2)", "Add Debit in Bank\n(Annexure 3)",
        "Less Credit in Bank\n(Annexure 4)", "GL Balance as per\nBank Recon (A)",
        "Balance as per\nMNP Ledger in Actual (B)", "Difference\n(B-A)"
    ]
    header_row = 6
    for i, h in enumerate(headers):
        cell = ws.cell(row=header_row, column=2 + i, value=h)
        cell.fill = navy_fill
        cell.font = navy_font

    totals = [0.0] * 7  # bank_bal, add_dr_mp, less_cr_mp, add_dr_bank, less_cr_bank, gl_bal, mp_bal

    for i, (bank_name, code, bank_key) in enumerate(FULL_ACCOUNTS):
        row = header_row + 1 + i
        ws.cell(row=row, column=2, value=bank_name)
        ws.cell(row=row, column=3, value=depot_code)
        ws.cell(row=row, column=4, value=code)

        add_dr_mp = 0
        add_dr_bank = 0

        if bank_key and bank_key in results_by_bank:
            r = results_by_bank[bank_key]
            unmatched_dr = r["ledger_debits"][r["ledger_debits"]["Status"] == "UNMATCHED"]
            add_dr_mp = unmatched_dr["Entered Amount DR"].sum()
            add_dr_bank = r["unmatched_bank_items"][BANKS[bank_key]["amount_col"]].sum()

        bank_bal = 0
        less_cr_mp = 0
        less_cr_bank = 0
        gl_bal = bank_bal + add_dr_mp - less_cr_mp + add_dr_bank - less_cr_bank
        mp_bal = gl_bal  # ledger actual balance; we don't have trial balance input, mirror GL bal
        diff = mp_bal - gl_bal

        vals = [bank_bal, add_dr_mp, less_cr_mp, add_dr_bank, less_cr_bank, gl_bal, mp_bal, diff]
        for j, v in enumerate(vals):
            ws.cell(row=row, column=5 + j, value=v)

        for k in range(6):
            totals[k] += vals[k] if k < len(vals) else 0

    total_row = header_row + 1 + len(FULL_ACCOUNTS)
    ws.cell(row=total_row, column=2, value="TOTAL").font = bold
    for j in range(8):
        col_vals = [
            ws.cell(row=header_row + 1 + i, column=5 + j).value or 0
            for i in range(len(FULL_ACCOUNTS))
        ]
        cell = ws.cell(row=total_row, column=5 + j, value=sum(col_vals))
        cell.font = bold

    for col_letter, width in zip("BCDEFGHIJKL", [26, 10, 10, 14, 16, 16, 14, 14, 16, 18, 12]):
        ws.column_dimensions[col_letter].width = width

    # ---------------- ANNEX 1 (Debit M&P) ----------------
    ws1 = wb.create_sheet("Annex 1 (Debit M&P)")
    ws1["A1"] = "Annexure 1"
    ws1["A2"] = "Debit in M&P Ledger - Collection Accounts"

    row_cursor = 4
    for bank_name, code, bank_key in FULL_ACCOUNTS:
        ws1.cell(row=row_cursor, column=1,
                 value=f"{bank_name} - 01.000.000.{depot_code}.000.{code}").font = bold
        row_cursor += 1

        for i, h in enumerate(ANNEX_LEFT_HEADERS):
            cell = ws1.cell(row=row_cursor, column=1 + i, value=h)
            cell.fill = navy_fill
            cell.font = navy_font
        row_cursor += 1

        items = []
        if bank_key and bank_key in results_by_bank:
            r = results_by_bank[bank_key]
            unmatched_dr = r["ledger_debits"][r["ledger_debits"]["Status"] == "UNMATCHED"]
            items = unmatched_dr.to_dict("records")

        if not items:
            for blank_row in range(2):
                ws1.cell(row=row_cursor, column=1, value=blank_row + 1)
                ws1.cell(row=row_cursor, column=7, value=0)
                row_cursor += 1
        else:
            for s_no, item in enumerate(items, start=1):
                ws1.cell(row=row_cursor, column=1, value=s_no)
                ws1.cell(row=row_cursor, column=3, value=str(item.get("Journal Line Description", "")))
                ws1.cell(row=row_cursor, column=7, value=item.get("Entered Amount DR"))
                ws1.cell(row=row_cursor, column=9,
                         value="Not matched to bank MIS this month - review manually")
                for c in range(1, 10):
                    ws1.cell(row=row_cursor, column=c).fill = red_fill
                row_cursor += 1

        row_cursor += 2  # gap between blocks

    for col_letter, width in zip("ABCDEFGHI", [8, 12, 18, 18, 9, 14, 14, 14, 44]):
        ws1.column_dimensions[col_letter].width = width

    # ---------------- ANNEX 3 (Debit Bank) ----------------
    ws3 = wb.create_sheet("Annex 3 (Debit Bank)")
    ws3["A1"] = "Annexure 3"
    ws3["A2"] = "Debit in Bank - Collection Accounts"

    row_cursor = 4
    for bank_name, code, bank_key in FULL_ACCOUNTS:
        ws3.cell(row=row_cursor, column=1,
                 value=f"{bank_name} - 01.000.000.{depot_code}.000.{code}").font = bold
        row_cursor += 1

        for i, h in enumerate(ANNEX_LEFT_HEADERS):
            cell = ws3.cell(row=row_cursor, column=1 + i, value=h)
            cell.fill = navy_fill
            cell.font = navy_font
        row_cursor += 1

        bank_items = []
        if bank_key and bank_key in results_by_bank:
            r = results_by_bank[bank_key]
            bank_items = r["unmatched_bank_items"].to_dict("records")

        if not bank_items:
            for blank_row in range(2):
                ws3.cell(row=row_cursor, column=1, value=blank_row + 1)
                ws3.cell(row=row_cursor, column=7, value=0)
                row_cursor += 1
        else:
            cfg = BANKS[bank_key]
            for s_no, item in enumerate(bank_items, start=1):
                ws3.cell(row=row_cursor, column=1, value=s_no)
                ws3.cell(row=row_cursor, column=2, value=str(item.get(cfg["date_col"], "")))
                ws3.cell(row=row_cursor, column=4, value=str(item.get(cfg["slip_col"], "")))
                ws3.cell(row=row_cursor, column=7, value=item.get(cfg["amount_col"]))
                ws3.cell(row=row_cursor, column=9,
                         value="In bank for this depot, not found in ledger this month - review manually")
                for c in range(1, 10):
                    ws3.cell(row=row_cursor, column=c).fill = red_fill
                row_cursor += 1

        row_cursor += 2

    for col_letter, width in zip("ABCDEFGHI", [8, 12, 18, 18, 9, 14, 14, 14, 44]):
        ws3.column_dimensions[col_letter].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def build_excel(results, depot_code):
    output = io.BytesIO()

    red_fill = PatternFill(start_color="FDEAEA", end_color="FDEAEA", fill_type="solid")
    green_fill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
    amber_fill = PatternFill(start_color="FEF3E0", end_color="FEF3E0", fill_type="solid")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:

        # ---------------- SUMMARY SHEET ----------------
        summary_rows = []
        for r in results:
            total = len(r["ledger_debits"])
            matched = (r["ledger_debits"]["Status"] == "MATCHED").sum()
            summary_rows.append({
                "Bank": r["account_label"],
                "Account Code": r["account_code"],
                "Ledger Debit Items": total,
                "Matched": matched,
                "Needs Review (Ledger)": total - matched,
                "Unmatched Bank Items": len(r["unmatched_bank_items"]),
            })
        summary_df = pd.DataFrame(summary_rows)
        summary_df.to_excel(writer, sheet_name="Summary", index=False, startrow=1)
        ws = writer.sheets["Summary"]
        ws.cell(row=1, column=1).value = f"Bank Reconciliation Summary - Depot {depot_code}"
        for col in range(1, len(summary_df.columns) + 1):
            ws.cell(row=2, column=col).fill = header_fill
            ws.cell(row=2, column=col).font = header_font

        # ---------------- ONE SHEET PER BANK ----------------
        for r in results:
            sheet_name = r["bank"][:31]

            dr_cols = [
                "Journal Line Number", "Journal Line Description",
                "Entered Amount DR", "Status", "Bank Deposit Slip",
                "Bank Amount", "Bank Date"
            ]
            dr_display = r["ledger_debits"][[c for c in dr_cols if c in r["ledger_debits"].columns]]

            dr_display.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)
            ws = writer.sheets[sheet_name]
            ws.cell(row=1, column=1).value = (
                f"{r['account_label']} ({r['account_code']}) - Ledger Debits vs Bank"
            )

            for col in range(1, len(dr_display.columns) + 1):
                ws.cell(row=3, column=col).fill = header_fill
                ws.cell(row=3, column=col).font = header_font

            status_col_idx = list(dr_display.columns).index("Status") + 1 if "Status" in dr_display.columns else None
            if status_col_idx:
                for row_num in range(4, 4 + len(dr_display)):
                    status_val = ws.cell(row=row_num, column=status_col_idx).value
                    fill = red_fill if status_val == "UNMATCHED" else green_fill
                    for col_num in range(1, len(dr_display.columns) + 1):
                        ws.cell(row=row_num, column=col_num).fill = fill

            # ---- unmatched bank items, appended below ----
            next_row = 4 + len(dr_display) + 2
            ws.cell(row=next_row, column=1).value = "Bank items for this depot with no matching ledger entry (review):"
            ws.cell(row=next_row, column=1).font = Font(bold=True)

            bank_items = r["unmatched_bank_items"]
            if len(bank_items) > 0:
                start = next_row + 1
                for col_num, col_name in enumerate(bank_items.columns, start=1):
                    cell = ws.cell(row=start, column=col_num)
                    cell.value = col_name
                    cell.fill = header_fill
                    cell.font = header_font
                for r_idx, (_, row_data) in enumerate(bank_items.iterrows(), start=start + 1):
                    for col_num, val in enumerate(row_data, start=1):
                        cell = ws.cell(row=r_idx, column=col_num)
                        cell.value = val
                        cell.fill = amber_fill

        # ---------------- CREDIT ENTRIES (reference only) ----------------
        all_credits = pd.concat(
            [r["ledger_credits"].assign(Bank=r["bank"]) for r in results if len(r["ledger_credits"]) > 0],
            ignore_index=True
        ) if any(len(r["ledger_credits"]) > 0 for r in results) else pd.DataFrame()

        if not all_credits.empty:
            cr_cols = ["Bank", "Journal Line Number", "Journal Line Description", "Entered Amount CR"]
            cr_display = all_credits[[c for c in cr_cols if c in all_credits.columns]]
            cr_display.to_excel(writer, sheet_name="Ledger Credits (ref)", index=False, startrow=1)
            ws = writer.sheets["Ledger Credits (ref)"]
            ws.cell(row=1, column=1).value = "Credit entries in ledger (bulk transfers-out, reference only)"

    output.seek(0)
    return output


# ---------------- UI ----------------

st.title("🏦 Bank Reconciliation Tool")
st.caption("Muller & Phipps Pakistan — matches ledger (GL) collections against bank MIS files by amount, flags exceptions for manual review")

st.divider()

col1, col2 = st.columns([1, 2])

with col1:
    depot_code = st.selectbox("Select Depot Code", DEPOT_CODES)

st.subheader("1. Upload GL File")
gl_file = st.file_uploader("GL file for this depot (.xlsx)", type=["xlsx"], key="gl_upload")

st.subheader("2. Upload Bank MIS Files")
st.caption("Upload whichever bank files you have this month — combined files covering all depots are fine, the tool filters by depot code automatically.")

bank_files = {}
bcol1, bcol2, bcol3, bcol4 = st.columns(4)
with bcol1:
    bank_files["MCB"] = st.file_uploader("MCB", type=["xlsx"], key="mcb_upload")
with bcol2:
    bank_files["HBL"] = st.file_uploader("HBL", type=["xlsx"], key="hbl_upload")
with bcol3:
    bank_files["UBL"] = st.file_uploader("UBL", type=["xlsx"], key="ubl_upload")
with bcol4:
    bank_files["FAYSAL"] = st.file_uploader("Faysal", type=["xlsx"], key="faysal_upload")

st.divider()

if st.button("🔄 Run Reconciliation"):

    if gl_file is None:
        st.warning("Please upload the GL file first.")
    else:
        try:
            gl_df = load_gl(gl_file)
        except Exception as e:
            st.error(f"Could not read GL file: {e}")
            st.stop()

        results = []
        errors = []

        for bank_key, file in bank_files.items():
            if file is None:
                continue
            try:
                bank_df = load_bank(file, bank_key)
                result = reconcile_account(gl_df, bank_df, bank_key, depot_code)
                results.append(result)
            except Exception as e:
                errors.append(f"{bank_key}: {e}")

        for err in errors:
            st.error(f"Error processing {err}")

        if not results:
            st.warning("No bank files were successfully processed. Upload at least one bank file.")
        else:
            st.session_state["recon_results"] = results
            st.session_state["recon_depot"] = depot_code
            st.success(f"Reconciliation complete for depot {depot_code}")

if "recon_results" in st.session_state:

    results = st.session_state["recon_results"]
    depot_code = st.session_state["recon_depot"]

    st.subheader("Summary")

    summary_data = []
    for r in results:
        total = len(r["ledger_debits"])
        matched = (r["ledger_debits"]["Status"] == "MATCHED").sum()
        summary_data.append({
            "Bank": r["account_label"],
            "Ledger Items": total,
            "Matched": matched,
            "Needs Review": total - matched,
            "Unmatched Bank Items": len(r["unmatched_bank_items"]),
        })
    st.dataframe(pd.DataFrame(summary_data), use_container_width=True, hide_index=True)

    st.subheader("Details by Bank")

    for r in results:
        with st.expander(f"{r['account_label']} — {len(r['ledger_debits'])} ledger items"):

            dr_display = r["ledger_debits"][
                [c for c in ["Journal Line Number", "Journal Line Description", "Entered Amount DR",
                             "Status", "Bank Deposit Slip", "Bank Amount", "Bank Date"]
                 if c in r["ledger_debits"].columns]
            ]

            def highlight_status(row):
                if row["Status"] == "UNMATCHED":
                    return ["background-color: #fdeaea"] * len(row)
                return ["background-color: #e6f4ea"] * len(row)

            st.dataframe(
                dr_display.style.apply(highlight_status, axis=1),
                use_container_width=True, hide_index=True
            )

            if len(r["unmatched_bank_items"]) > 0:
                st.markdown("**Bank items for this depot with no matching ledger entry:**")
                st.dataframe(r["unmatched_bank_items"], use_container_width=True, hide_index=True)

    st.divider()

    results_by_bank = {r["bank"]: r for r in results}
    template_excel = build_template_excel(results_by_bank, depot_code)

    st.download_button(
        "⬇️ Download Reconciliation (Main Recon + Annex 1 + Annex 3 format)",
        data=template_excel,
        file_name=f"BR_Depot_{depot_code}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    simple_excel = build_excel(results, depot_code)

    st.download_button(
        "⬇️ Download Simple Summary (one sheet per bank, easier to skim)",
        data=simple_excel,
        file_name=f"Bank_Reconciliation_Depot_{depot_code}_Simple.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
