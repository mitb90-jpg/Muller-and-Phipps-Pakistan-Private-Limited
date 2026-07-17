import streamlit as st
import pandas as pd
import io
from openpyxl.styles import PatternFill, Font

# ---------------- PAGE CONFIG ----------------
st.set_page_config(
    page_title="Bank Reconciliation Tool - Muller & Phipps",
    page_icon="🏦",
    layout="wide"
)

# ---------------- BANK CONFIGURATION ----------------
# To add a new bank later: add an entry here with its sheet name, header
# row (0-indexed), and the column names for deposit slip / amount / depot.
BANKS = {
    "MCB": {
        "account_code": 212201, "account_label": "MCB Bank Ltd",
        "sheet": "MCB", "header": 1,
        "slip_col": "Deposit Slip No.", "amount_col": "Amount",
        "depot_col": "Dealer Code", "date_col": "Credit Date",
    },
    "HBL": {
        "account_code": 212202, "account_label": "Habib Bank Limited",
        "sheet": "HBL", "header": 6,
        "slip_col": "Deposit Slip No.", "amount_col": "Amount",
        "depot_col": "Dealer Code", "date_col": "Credit Date",
    },
    "UBL": {
        "account_code": 212205, "account_label": "United Bank Limited",
        "sheet": "UBL", "header": 0,
        "slip_col": "DEPOSIT SLIP NO", "amount_col": "AMOUNT",
        "depot_col": "DEPOT'S CODE", "date_col": "REALIZATION DATE",
    },
    "FAYSAL": {
        "account_code": 212209, "account_label": "Faysal Bank Limited",
        "sheet": "MIS", "header": 4,
        "slip_col": "Deposit No.", "amount_col": "Amount",
        "depot_col": "Dealer Code", "date_col": "Credit Date",
    },
}

DEPOT_CODES = [612, 613, 614, 651, 657, 645, 646, 733, 766, 778]
TOLERANCE = 1.0  # rupees - absorbs rounding differences


# ---------------- CSS ----------------
st.markdown("""
<style>
div.stDownloadButton > button, div.stButton > button {
    width: 100%;
    background-color: #1f4e79;
    color: white;
    font-size: 16px;
    font-weight: bold;
    padding: 12px;
    border-radius: 10px;
    border: none;
}
div.stDownloadButton > button:hover, div.stButton > button:hover {
    background-color: #163a5c;
}
</style>
""", unsafe_allow_html=True)


# ---------------- HELPERS ----------------

def load_gl(file):
    df = pd.read_excel(file, sheet_name="GL")
    df.columns = [str(c).strip() for c in df.columns]
    return df


def load_bank(file, bank_key):
    cfg = BANKS[bank_key]
    df = pd.read_excel(file, sheet_name=cfg["sheet"], header=cfg["header"])
    df.columns = [str(c).strip() for c in df.columns]
    return df


def reconcile_account(gl_df, bank_df, bank_key, depot_code):
    cfg = BANKS[bank_key]

    gl_sub = gl_df[gl_df["Natural Account Segment"] == cfg["account_code"]].copy()
    dr = gl_sub[gl_sub["Entered Amount DR"].notna()].copy()
    cr = gl_sub[gl_sub["Entered Amount CR"].notna()].copy()

    bank_sub = bank_df[bank_df[cfg["depot_col"]] == depot_code].copy()
    bank_sub["_used"] = False

    dr["Status"] = "UNMATCHED"
    dr["Bank Deposit Slip"] = pd.Series([""] * len(dr), index=dr.index, dtype="object")
    dr["Bank Amount"] = None
    dr["Bank Date"] = pd.Series([""] * len(dr), index=dr.index, dtype="object")

    for idx, row in dr.iterrows():
        amt = row["Entered Amount DR"]
        if pd.isna(amt):
            continue
        candidates = bank_sub[
            (~bank_sub["_used"]) & ((bank_sub[cfg["amount_col"]] - amt).abs() <= TOLERANCE)
        ]
        if len(candidates) > 0:
            match_idx = candidates.index[0]
            bank_sub.loc[match_idx, "_used"] = True
            dr.loc[idx, "Status"] = "MATCHED"
            dr.loc[idx, "Bank Deposit Slip"] = str(bank_sub.loc[match_idx, cfg["slip_col"]])
            dr.loc[idx, "Bank Amount"] = bank_sub.loc[match_idx, cfg["amount_col"]]
            dr.loc[idx, "Bank Date"] = str(bank_sub.loc[match_idx, cfg["date_col"]])

    unmatched_bank = bank_sub[~bank_sub["_used"]].drop(columns=["_used"])

    return {
        "bank": bank_key,
        "account_code": cfg["account_code"],
        "account_label": cfg["account_label"],
        "ledger_debits": dr,
        "ledger_credits": cr,
        "unmatched_bank_items": unmatched_bank,
    }


# Full list of GLR accounts in the exact order used in the real template.
# bank_key links to BANKS dict where we have a parser; None = no data source yet.
FULL_ACCOUNTS = [
    ("MCB Bank Ltd", 212201, "MCB"),
    ("Habib Bank Limited", 212202, "HBL"),
    ("National Bank of Pakistan", 212203, None),
    ("HBL Konnect", 212204, None),
    ("United Bank Limited", 212205, "UBL"),
    ("Bank Al Falah Limited", 212206, None),
    ("Bank Al Habib Limited", 212207, None),
    ("Samba Bank Limited", 212208, None),
    ("Faysal Bank Limited", 212209, "FAYSAL"),
    ("Dubai Islamic Bank Limited", 212210, None),
    ("Meezan Bank Limited", 212211, None),
    ("Habib Metropoliton Bank", 212212, None),
    ("Meezan Bank Limited (MFS)", 212213, None),
    ("Habib Bank Limited (MFS)", 212214, None),
    ("Askari Bank Limited", 212215, None),
    ("Bank of Punjab", 212216, None),
    ("United Bank Limited (MFS)", 212217, None),
    ("Bank Islami Limited", 212218, None),
    ("UBL Omni", 212219, None),
    ("Standard Chartered Bank", 212503, None),
]

ANNEX_LEFT_HEADERS = [
    "S. No.", "Dates", "D/Slip No\nas Per Salesman", "D/Slip No\nas Per\nBank MIS",
    "Ch/Cash", "Chq. No.", "Amount", "Clearance Date", "Remarks"
]


IMPREST_ACCOUNTS = [
    ("MCB Bank Ltd", 212401),
    ("Habib Bank Limited", 212402),
    ("National Bank of Pakistan", 212404),
    ("Standard Chartered Bank", 212405),
    ("Bank Al Falah Limited", 212407),
    ("Bank Al Habib Limited", 212408),
    ("Dubai Islamic Bank Limited", 212409),
    ("Faysal Bank Limited", 212410),
]


def build_template_excel(results_by_bank, depot_code, depot_name=""):
    """Builds output matching the real Main Reconciliation / Annex 1 (Debit M&P) /
    Annex 3 (Debit Bank) template: same formatting (wrapped yellow headers, thin
    borders, merged titles, column widths) and the same live formulas linking
    Main Reconciliation totals to each Annex sheet's block SUM."""

    import openpyxl
    from openpyxl.styles import Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    HEADER_FILL = PatternFill(start_color="FFFCD5", end_color="FFFCD5", fill_type="solid")
    RED_FILL = PatternFill(start_color="FDEAEA", end_color="FDEAEA", fill_type="solid")
    BOLD = Font(bold=True, size=10)
    NORMAL = Font(size=10)
    WRAP_CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")
    THIN = Side(style="thin")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def style_header_cell(cell, text):
        cell.value = text
        cell.font = BOLD
        cell.fill = HEADER_FILL
        cell.alignment = WRAP_CENTER
        cell.border = BORDER

    def style_data_cell(cell, value=None, bold=False, red=False):
        if value is not None:
            cell.value = value
        cell.font = BOLD if bold else NORMAL
        cell.border = BORDER
        if red:
            cell.fill = RED_FILL

    # ---------------- ANNEX 1 (Debit M&P) ----------------
    ws1 = wb.active
    ws1.title = "Annex 1 (Debit M&P)"
    ws1["A1"] = "Annexure 1"
    ws1["A1"].font = BOLD
    ws1["A2"] = "Debit in M&P Ledger - Collection Accounts"
    ws1["A2"].font = BOLD

    annex1_totals = {}  # code -> (sheet_name, total_row)
    row_cursor = 4
    for bank_name, code, bank_key in FULL_ACCOUNTS:
        ws1.cell(row=row_cursor, column=1,
                 value=f"{bank_name} - 01.000.000.{depot_code}.000.{code}").font = BOLD
        row_cursor += 1

        for i, h in enumerate(ANNEX_LEFT_HEADERS):
            style_header_cell(ws1.cell(row=row_cursor, column=1 + i), h)
        header_row_here = row_cursor
        ws1.row_dimensions[header_row_here].height = 40
        row_cursor += 1

        items = []
        if bank_key and bank_key in results_by_bank:
            r = results_by_bank[bank_key]
            unmatched_dr = r["ledger_debits"][r["ledger_debits"]["Status"] == "UNMATCHED"]
            items = unmatched_dr.to_dict("records")

        first_item_row = row_cursor
        n_rows = max(len(items), 2)
        for i in range(n_rows):
            if i < len(items):
                item = items[i]
                style_data_cell(ws1.cell(row=row_cursor, column=1), i + 1)
                style_data_cell(ws1.cell(row=row_cursor, column=3), str(item.get("Journal Line Description", "")))
                style_data_cell(ws1.cell(row=row_cursor, column=7), item.get("Entered Amount DR"), red=True)
                style_data_cell(ws1.cell(row=row_cursor, column=9), "Not matched to bank MIS this month - review manually", red=True)
                for c in [2, 4, 5, 6, 8]:
                    style_data_cell(ws1.cell(row=row_cursor, column=c), red=True)
            else:
                style_data_cell(ws1.cell(row=row_cursor, column=1), i + 1)
                style_data_cell(ws1.cell(row=row_cursor, column=7), 0)
                for c in [2, 3, 4, 5, 6, 8, 9]:
                    style_data_cell(ws1.cell(row=row_cursor, column=c))
            row_cursor += 1
        last_item_row = row_cursor - 1

        total_row = row_cursor
        style_data_cell(ws1.cell(row=total_row, column=1), "Total", bold=True)
        for c in [2, 3, 4, 5, 6, 8, 9]:
            style_data_cell(ws1.cell(row=total_row, column=c), bold=True)
        total_cell = ws1.cell(row=total_row, column=7)
        total_cell.value = f"=SUM(G{first_item_row}:G{last_item_row})"
        total_cell.font = BOLD
        total_cell.border = BORDER
        annex1_totals[code] = total_row
        row_cursor += 2

    for col_letter, width in zip("ABCDEFGHI", [8, 12, 20, 18, 9, 12, 14, 14, 44]):
        ws1.column_dimensions[col_letter].width = width

    # ---------------- ANNEX 3 (Debit Bank) ----------------
    ws3 = wb.create_sheet("Annex 3 (Debit Bank)")
    ws3["A1"] = "Annexure 3"
    ws3["A1"].font = BOLD
    ws3["A2"] = "Debit in Bank - Collection Accounts"
    ws3["A2"].font = BOLD

    annex3_totals = {}
    row_cursor = 4
    for bank_name, code, bank_key in FULL_ACCOUNTS:
        ws3.cell(row=row_cursor, column=1,
                 value=f"{bank_name} - 01.000.000.{depot_code}.000.{code}").font = BOLD
        row_cursor += 1

        for i, h in enumerate(ANNEX_LEFT_HEADERS):
            style_header_cell(ws3.cell(row=row_cursor, column=1 + i), h)
        ws3.row_dimensions[row_cursor].height = 40
        row_cursor += 1

        bank_items = []
        cfg = BANKS.get(bank_key) if bank_key else None
        if bank_key and bank_key in results_by_bank:
            r = results_by_bank[bank_key]
            bank_items = r["unmatched_bank_items"].to_dict("records")

        first_item_row = row_cursor
        n_rows = max(len(bank_items), 2)
        for i in range(n_rows):
            if i < len(bank_items):
                item = bank_items[i]
                style_data_cell(ws3.cell(row=row_cursor, column=1), i + 1)
                style_data_cell(ws3.cell(row=row_cursor, column=2), str(item.get(cfg["date_col"], "")))
                style_data_cell(ws3.cell(row=row_cursor, column=4), str(item.get(cfg["slip_col"], "")))
                style_data_cell(ws3.cell(row=row_cursor, column=7), item.get(cfg["amount_col"]), red=True)
                style_data_cell(ws3.cell(row=row_cursor, column=9), "In bank for this depot, not found in ledger this month - review manually", red=True)
                for c in [3, 5, 6, 8]:
                    style_data_cell(ws3.cell(row=row_cursor, column=c), red=True)
            else:
                style_data_cell(ws3.cell(row=row_cursor, column=1), i + 1)
                style_data_cell(ws3.cell(row=row_cursor, column=7), 0)
                for c in [2, 3, 4, 5, 6, 8, 9]:
                    style_data_cell(ws3.cell(row=row_cursor, column=c))
            row_cursor += 1
        last_item_row = row_cursor - 1

        total_row = row_cursor
        style_data_cell(ws3.cell(row=total_row, column=1), "Total", bold=True)
        for c in [2, 3, 4, 5, 6, 8, 9]:
            style_data_cell(ws3.cell(row=total_row, column=c), bold=True)
        total_cell = ws3.cell(row=total_row, column=7)
        total_cell.value = f"=SUM(G{first_item_row}:G{last_item_row})"
        total_cell.font = BOLD
        total_cell.border = BORDER
        annex3_totals[code] = total_row
        row_cursor += 2

    for col_letter, width in zip("ABCDEFGHI", [8, 12, 20, 18, 9, 12, 14, 14, 44]):
        ws3.column_dimensions[col_letter].width = width

    # ---------------- ANNEX 2 (Credit M&P) ----------------
    # No credit-side matching logic yet (these are bulk monthly transfer-out
    # entries, not per-transaction) - list ledger credit entries for reference
    # and total them, so Main Reconciliation can still link to a real total.
    ws2 = wb.create_sheet("Annex 2 (Credit M&P)")
    ws2["A1"] = "Annexure 2"
    ws2["A1"].font = BOLD
    ws2["A2"] = "Credit in M&P Ledger - Collection Accounts"
    ws2["A2"].font = BOLD

    annex2_totals = {}
    row_cursor = 4
    for bank_name, code, bank_key in FULL_ACCOUNTS:
        ws2.cell(row=row_cursor, column=1,
                 value=f"{bank_name} - 01.000.000.{depot_code}.000.{code}").font = BOLD
        row_cursor += 1
        for i, h in enumerate(ANNEX_LEFT_HEADERS):
            style_header_cell(ws2.cell(row=row_cursor, column=1 + i), h)
        ws2.row_dimensions[row_cursor].height = 40
        row_cursor += 1

        items = []
        if bank_key and bank_key in results_by_bank:
            r = results_by_bank[bank_key]
            items = r["ledger_credits"].to_dict("records")

        first_item_row = row_cursor
        n_rows = max(len(items), 2)
        for i in range(n_rows):
            if i < len(items):
                item = items[i]
                style_data_cell(ws2.cell(row=row_cursor, column=1), i + 1)
                style_data_cell(ws2.cell(row=row_cursor, column=3), str(item.get("Journal Line Description", "")))
                style_data_cell(ws2.cell(row=row_cursor, column=7), item.get("Entered Amount CR"))
                style_data_cell(ws2.cell(row=row_cursor, column=9), "Bulk monthly transfer-out per ledger (reference only)")
                for c in [2, 4, 5, 6, 8]:
                    style_data_cell(ws2.cell(row=row_cursor, column=c))
            else:
                style_data_cell(ws2.cell(row=row_cursor, column=1), i + 1)
                style_data_cell(ws2.cell(row=row_cursor, column=7), 0)
                for c in [2, 3, 4, 5, 6, 8, 9]:
                    style_data_cell(ws2.cell(row=row_cursor, column=c))
            row_cursor += 1
        last_item_row = row_cursor - 1

        total_row = row_cursor
        style_data_cell(ws2.cell(row=total_row, column=1), "Total", bold=True)
        for c in [2, 3, 4, 5, 6, 8, 9]:
            style_data_cell(ws2.cell(row=total_row, column=c), bold=True)
        tc = ws2.cell(row=total_row, column=7)
        tc.value = f"=SUM(G{first_item_row}:G{last_item_row})"
        tc.font = BOLD
        tc.border = BORDER
        annex2_totals[code] = total_row
        row_cursor += 2

    for col_letter, width in zip("ABCDEFGHI", [8, 12, 20, 18, 9, 12, 14, 14, 44]):
        ws2.column_dimensions[col_letter].width = width

    # ---------------- ANNEX 4 (Credit Bank) ----------------
    # No data source (would require full bank statements, not collection MIS
    # listings) - skeleton only, matching template's own empty state.
    ws4 = wb.create_sheet("Annex 4 (Credit Bank)")
    ws4["A1"] = "Annexure 4"
    ws4["A1"].font = BOLD
    ws4["A2"] = "Credit in Bank - Collection Accounts"
    ws4["A2"].font = BOLD

    annex4_totals = {}
    row_cursor = 4
    for bank_name, code, bank_key in FULL_ACCOUNTS:
        ws4.cell(row=row_cursor, column=1,
                 value=f"{bank_name} - 01.000.000.{depot_code}.000.{code}").font = BOLD
        row_cursor += 1
        for i, h in enumerate(ANNEX_LEFT_HEADERS):
            style_header_cell(ws4.cell(row=row_cursor, column=1 + i), h)
        ws4.row_dimensions[row_cursor].height = 40
        row_cursor += 1

        first_item_row = row_cursor
        for i in range(2):
            style_data_cell(ws4.cell(row=row_cursor, column=1), i + 1)
            style_data_cell(ws4.cell(row=row_cursor, column=7), 0)
            for c in [2, 3, 4, 5, 6, 8, 9]:
                style_data_cell(ws4.cell(row=row_cursor, column=c))
            row_cursor += 1
        last_item_row = row_cursor - 1

        total_row = row_cursor
        style_data_cell(ws4.cell(row=total_row, column=1), "Total", bold=True)
        for c in [2, 3, 4, 5, 6, 8, 9]:
            style_data_cell(ws4.cell(row=total_row, column=c), bold=True)
        tc = ws4.cell(row=total_row, column=7)
        tc.value = f"=SUM(G{first_item_row}:G{last_item_row})"
        tc.font = BOLD
        tc.border = BORDER
        annex4_totals[code] = total_row
        row_cursor += 2

    for col_letter, width in zip("ABCDEFGHI", [8, 12, 20, 18, 9, 12, 14, 14, 44]):
        ws4.column_dimensions[col_letter].width = width

    # ---------------- IMPREST ANNEX 1-4 (no data source - skeleton only) ----------------
    imprest_totals = {1: {}, 2: {}, 3: {}, 4: {}}
    imprest_titles = {
        1: ("Annex 1 (Debit M&P) - Imprest", "Debit in M&P Ledger (Imprest Accounts)"),
        2: ("Annex 2 (Credit M&P) - Imprest", "Credit in M&P Ledger (Imprest Accounts)"),
        3: ("Annex 3 (Debit Bank) - Imprest", "Debit in Bank (Imprest Accounts)"),
        4: ("Annex 4 (Credit Bank) - Imprest", "Credit in Bank (Imprest Accounts)"),
    }
    for annex_num in [1, 2, 3, 4]:
        sheet_title, subtitle = imprest_titles[annex_num]
        wsx = wb.create_sheet(sheet_title)
        wsx["A1"] = "Annexure " + str(annex_num)
        wsx["A1"].font = BOLD
        wsx["A2"] = subtitle
        wsx["A2"].font = BOLD

        row_cursor = 4
        for bank_name, code in IMPREST_ACCOUNTS:
            wsx.cell(row=row_cursor, column=1,
                     value=f"{bank_name} - 01.000.000.{depot_code}.000.{code}").font = BOLD
            row_cursor += 1
            for i, h in enumerate(ANNEX_LEFT_HEADERS[:8]):  # Imprest has no "Ch/Cash" split in sample
                style_header_cell(wsx.cell(row=row_cursor, column=1 + i), h)
            row_cursor += 1

            first_item_row = row_cursor
            for i in range(2):
                style_data_cell(wsx.cell(row=row_cursor, column=1), i + 1)
                style_data_cell(wsx.cell(row=row_cursor, column=6), 0)
                for c in [2, 3, 4, 5, 7, 8]:
                    style_data_cell(wsx.cell(row=row_cursor, column=c))
                row_cursor += 1
            last_item_row = row_cursor - 1

            total_row = row_cursor
            style_data_cell(wsx.cell(row=total_row, column=1), "Total", bold=True)
            tc = wsx.cell(row=total_row, column=6)
            tc.value = f"=SUM(F{first_item_row}:F{last_item_row})"
            tc.font = BOLD
            tc.border = BORDER
            imprest_totals[annex_num][code] = total_row
            row_cursor += 2

        for col_letter, width in zip("ABCDEFGH", [8, 12, 18, 9, 12, 14, 14, 44]):
            wsx.column_dimensions[col_letter].width = width

    # ---------------- CASH IN HAND SHEETS (no digital data source) ----------------
    for sheet_name, label in [
        ("212101 (CIH - Sale Proceed)", "Cash in Hand - Sale Proceed"),
        ("212102 (CIH - Petty Cash)", "Cash in Hand - Petty Cash"),
    ]:
        wsc = wb.create_sheet(sheet_name)
        wsc["A1"] = "Muller & Phipps Pakistan (Pvt.) Ltd."
        wsc["A1"].font = BOLD
        wsc["A2"] = f"Physical Cash Count :: {label}"
        wsc["A3"] = (
            "This is a physical cash count performed manually in the branch - "
            "there is no digital file this tool can generate it from. "
            "Please fill this sheet in manually as before."
        )
        wsc["A3"].font = Font(italic=True, size=9)
        wsc.column_dimensions["A"].width = 70

    # ---------------- CODE SHEET (static reference, matches original) ----------------
    wscode = wb.create_sheet("Code")
    code_rows = [("Cash", n, 209, c) for n, c in [
        ("Cash in Hand - Sale Proceed", 212101), ("Cash in Hand - Petty Cash", 212102)
    ]]
    for bank_name, code, _ in FULL_ACCOUNTS:
        code_rows.append(("Cash at Bank", bank_name, 209, code))
    for i, row_vals in enumerate(code_rows, start=1):
        for j, v in enumerate(row_vals, start=1):
            wscode.cell(row=i, column=j, value=v)
    for col_letter, width in zip("ABCD", [16, 30, 8, 10]):
        wscode.column_dimensions[col_letter].width = width

    # ---------------- MAIN RECONCILIATION ----------------
    ws = wb.create_sheet("Main Reconciliation", 0)

    ws.merge_cells("B1:L1")
    ws.merge_cells("B2:L2")
    ws.merge_cells("B3:L3")
    ws["B2"] = "Muller & Phipps Pakistan (Pvt) Ltd"
    ws["B2"].font = BOLD
    ws["B3"] = "Reconciliation of CMD / Collection Accounts"
    ws["B3"].font = BOLD
    ws["B5"] = "Depot Name:"
    ws["B5"].font = BOLD
    ws["C5"] = depot_code
    ws["C5"].font = BOLD
    ws["E5"] = depot_name
    ws["E5"].font = BOLD

    headers = [
        "Branch Name \nAs Per GLR", "Location \nCode", "GLR \nCode No.",
        "Collection Bank Balance", "Add \nDebit in \nM&P Ledger \n(Annexure 1)",
        "Less\nCredit in \nM&P Ledger \n(Annexure 2)", "Add\nDebit in Bank \n(Annexure 3)",
        "Less\nCredit in Bank\n(Annexure 4)", "GL Balance \nas per\n Bank Recon \n(A)",
        "Balance \nas per\nMNP Ledger\nin Actual (B)*", "Difference \n(B-A)"
    ]
    header_row = 6
    for i, h in enumerate(headers):
        style_header_cell(ws.cell(row=header_row, column=2 + i), h)
    ws.row_dimensions[header_row].height = 66.75

    for i, (bank_name, code, bank_key) in enumerate(FULL_ACCOUNTS):
        row = header_row + 1 + i
        style_data_cell(ws.cell(row=row, column=2), bank_name)
        style_data_cell(ws.cell(row=row, column=3), depot_code)
        style_data_cell(ws.cell(row=row, column=4), code)
        style_data_cell(ws.cell(row=row, column=5), 0, bold=True)  # Collection Bank Balance - fill manually

        if code in annex1_totals:
            ws.cell(row=row, column=6).value = f"='Annex 1 (Debit M&P)'!G{annex1_totals[code]}"
        else:
            ws.cell(row=row, column=6).value = 0
        style_data_cell(ws.cell(row=row, column=6))

        if code in annex2_totals:
            ws.cell(row=row, column=7).value = f"='Annex 2 (Credit M&P)'!G{annex2_totals[code]}"
        else:
            ws.cell(row=row, column=7).value = 0
        style_data_cell(ws.cell(row=row, column=7))

        if code in annex3_totals:
            ws.cell(row=row, column=8).value = f"='Annex 3 (Debit Bank)'!G{annex3_totals[code]}"
        else:
            ws.cell(row=row, column=8).value = 0
        style_data_cell(ws.cell(row=row, column=8))

        if code in annex4_totals:
            ws.cell(row=row, column=9).value = f"='Annex 4 (Credit Bank)'!G{annex4_totals[code]}"
        else:
            ws.cell(row=row, column=9).value = 0
        style_data_cell(ws.cell(row=row, column=9))

        gl_bal_cell = ws.cell(row=row, column=10)
        gl_bal_cell.value = f"=SUM(E{row}:I{row})"
        gl_bal_cell.font = BOLD
        gl_bal_cell.border = BORDER

        style_data_cell(ws.cell(row=row, column=11), 0, bold=True)  # Balance per ledger - fill manually

        diff_cell = ws.cell(row=row, column=12)
        diff_cell.value = f"=K{row}-J{row}"
        diff_cell.font = BOLD
        diff_cell.border = BORDER

    total_row = header_row + 1 + len(FULL_ACCOUNTS)
    style_data_cell(ws.cell(row=total_row, column=2), "Total", bold=True)
    first_data_row = header_row + 1
    last_data_row = total_row - 1
    for col in range(5, 13):
        col_letter = get_column_letter(col)
        cell = ws.cell(row=total_row, column=col)
        cell.value = f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})"
        cell.font = BOLD
        cell.border = BORDER

    ws.cell(row=total_row + 2, column=2,
            value="* Fill 'Collection Bank Balance' and 'Balance as per MNP Ledger in Actual' manually from bank statement / trial balance.").font = Font(italic=True, size=9)

    # ---------------- IMPREST BANK ACCOUNTS TABLE ----------------
    imprest_header_row = total_row + 4
    ws.cell(row=imprest_header_row - 1, column=2, value="Reconciliation of Imprest Bank Accounts").font = BOLD
    for i, h in enumerate(headers):
        h2 = h.replace("Collection Bank Balance", "Balance as per\nBank Statement") \
             .replace("As Per GLR", "As Per GLR")
        style_header_cell(ws.cell(row=imprest_header_row, column=2 + i), h2)

    for i, (bank_name, code) in enumerate(IMPREST_ACCOUNTS):
        row = imprest_header_row + 1 + i
        style_data_cell(ws.cell(row=row, column=2), bank_name)
        style_data_cell(ws.cell(row=row, column=3), depot_code)
        style_data_cell(ws.cell(row=row, column=4), code)
        style_data_cell(ws.cell(row=row, column=5), 0, bold=True)

        ws.cell(row=row, column=6).value = f"='Annex 1 (Debit M&P) - Imprest'!F{imprest_totals[1].get(code, 4)}" if code in imprest_totals[1] else 0
        style_data_cell(ws.cell(row=row, column=6))
        ws.cell(row=row, column=7).value = f"='Annex 2 (Credit M&P) - Imprest'!F{imprest_totals[2].get(code, 4)}" if code in imprest_totals[2] else 0
        style_data_cell(ws.cell(row=row, column=7))
        ws.cell(row=row, column=8).value = f"='Annex 3 (Debit Bank) - Imprest'!F{imprest_totals[3].get(code, 4)}" if code in imprest_totals[3] else 0
        style_data_cell(ws.cell(row=row, column=8))
        ws.cell(row=row, column=9).value = f"='Annex 4 (Credit Bank) - Imprest'!F{imprest_totals[4].get(code, 4)}" if code in imprest_totals[4] else 0
        style_data_cell(ws.cell(row=row, column=9))

        gl_bal_cell = ws.cell(row=row, column=10)
        gl_bal_cell.value = f"=SUM(E{row}:I{row})"
        gl_bal_cell.font = BOLD
        gl_bal_cell.border = BORDER
        style_data_cell(ws.cell(row=row, column=11), 0, bold=True)
        diff_cell = ws.cell(row=row, column=12)
        diff_cell.value = f"=K{row}-J{row}"
        diff_cell.font = BOLD
        diff_cell.border = BORDER

    imprest_total_row = imprest_header_row + 1 + len(IMPREST_ACCOUNTS)
    style_data_cell(ws.cell(row=imprest_total_row, column=2), "Total", bold=True)
    for col in range(5, 13):
        col_letter = get_column_letter(col)
        cell = ws.cell(row=imprest_total_row, column=col)
        cell.value = f"=SUM({col_letter}{imprest_header_row + 1}:{col_letter}{imprest_total_row - 1})"
        cell.font = BOLD
        cell.border = BORDER

    # ---------------- CASH IN HAND (reference note only) ----------------
    cih_row = imprest_total_row + 3
    ws.cell(row=cih_row, column=2,
            value="Cash in Hand (Sale Proceed / Petty Cash) - see 'CIH' sheets. Physical count required, fill manually.").font = Font(italic=True, size=9)

    # ---------------- SIGNATURE BLOCK ----------------
    sig_row = cih_row + 3
    ws.cell(row=sig_row, column=2, value="_______________________").font = NORMAL
    ws.cell(row=sig_row, column=9, value="_______________________").font = NORMAL
    ws.cell(row=sig_row + 1, column=2, value="Prepared By").font = BOLD
    ws.cell(row=sig_row + 1, column=9, value="Reviewed By").font = BOLD

    col_widths = {"B": 25.5, "C": 8.9, "D": 9.3, "E": 13.1, "F": 15.1,
                  "G": 15.1, "H": 14.1, "I": 14.1, "J": 15.1, "K": 16, "L": 12}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def build_excel(results, depot_code):
    output = io.BytesIO()

    red_fill = PatternFill(start_color="FDEAEA", end_color="FDEAEA", fill_type="solid")
    green_fill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
    amber_fill = PatternFill(start_color="FEF3E0", end_color="FEF3E0", fill_type="solid")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:

        # ---------------- SUMMARY SHEET ----------------
        summary_rows = []
        for r in results:
            total = len(r["ledger_debits"])
            matched = (r["ledger_debits"]["Status"] == "MATCHED").sum()
            summary_rows.append({
                "Bank": r["account_label"],
                "Account Code": r["account_code"],
                "Ledger Debit Items": total,
                "Matched": matched,
                "Needs Review (Ledger)": total - matched,
                "Unmatched Bank Items": len(r["unmatched_bank_items"]),
            })
        summary_df = pd.DataFrame(summary_rows)
        summary_df.to_excel(writer, sheet_name="Summary", index=False, startrow=1)
        ws = writer.sheets["Summary"]
        ws.cell(row=1, column=1).value = f"Bank Reconciliation Summary - Depot {depot_code}"
        for col in range(1, len(summary_df.columns) + 1):
            ws.cell(row=2, column=col).fill = header_fill
            ws.cell(row=2, column=col).font = header_font

        # ---------------- ONE SHEET PER BANK ----------------
        for r in results:
            sheet_name = r["bank"][:31]

            dr_cols = [
                "Journal Line Number", "Journal Line Description",
                "Entered Amount DR", "Status", "Bank Deposit Slip",
                "Bank Amount", "Bank Date"
            ]
            dr_display = r["ledger_debits"][[c for c in dr_cols if c in r["ledger_debits"].columns]]

            dr_display.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)
            ws = writer.sheets[sheet_name]
            ws.cell(row=1, column=1).value = (
                f"{r['account_label']} ({r['account_code']}) - Ledger Debits vs Bank"
            )

            for col in range(1, len(dr_display.columns) + 1):
                ws.cell(row=3, column=col).fill = header_fill
                ws.cell(row=3, column=col).font = header_font

            status_col_idx = list(dr_display.columns).index("Status") + 1 if "Status" in dr_display.columns else None
            if status_col_idx:
                for row_num in range(4, 4 + len(dr_display)):
                    status_val = ws.cell(row=row_num, column=status_col_idx).value
                    fill = red_fill if status_val == "UNMATCHED" else green_fill
                    for col_num in range(1, len(dr_display.columns) + 1):
                        ws.cell(row=row_num, column=col_num).fill = fill

            # ---- unmatched bank items, appended below ----
            next_row = 4 + len(dr_display) + 2
            ws.cell(row=next_row, column=1).value = "Bank items for this depot with no matching ledger entry (review):"
            ws.cell(row=next_row, column=1).font = Font(bold=True)

            bank_items = r["unmatched_bank_items"]
            if len(bank_items) > 0:
                start = next_row + 1
                for col_num, col_name in enumerate(bank_items.columns, start=1):
                    cell = ws.cell(row=start, column=col_num)
                    cell.value = col_name
                    cell.fill = header_fill
                    cell.font = header_font
                for r_idx, (_, row_data) in enumerate(bank_items.iterrows(), start=start + 1):
                    for col_num, val in enumerate(row_data, start=1):
                        cell = ws.cell(row=r_idx, column=col_num)
                        cell.value = val
                        cell.fill = amber_fill

        # ---------------- CREDIT ENTRIES (reference only) ----------------
        all_credits = pd.concat(
            [r["ledger_credits"].assign(Bank=r["bank"]) for r in results if len(r["ledger_credits"]) > 0],
            ignore_index=True
        ) if any(len(r["ledger_credits"]) > 0 for r in results) else pd.DataFrame()

        if not all_credits.empty:
            cr_cols = ["Bank", "Journal Line Number", "Journal Line Description", "Entered Amount CR"]
            cr_display = all_credits[[c for c in cr_cols if c in all_credits.columns]]
            cr_display.to_excel(writer, sheet_name="Ledger Credits (ref)", index=False, startrow=1)
            ws = writer.sheets["Ledger Credits (ref)"]
            ws.cell(row=1, column=1).value = "Credit entries in ledger (bulk transfers-out, reference only)"

    output.seek(0)
    return output


# ---------------- UI ----------------

st.title("🏦 Bank Reconciliation Tool")
st.caption("Muller & Phipps Pakistan — matches ledger (GL) collections against bank MIS files by amount, flags exceptions for manual review")

st.divider()

col1, col2 = st.columns([1, 2])

with col1:
    depot_code = st.selectbox("Select Depot Code", DEPOT_CODES)

st.subheader("1. Upload GL File")
gl_file = st.file_uploader("GL file for this depot (.xlsx)", type=["xlsx"], key="gl_upload")

st.subheader("2. Upload Bank MIS Files")
st.caption("Upload whichever bank files you have this month — combined files covering all depots are fine, the tool filters by depot code automatically.")

bank_files = {}
bcol1, bcol2, bcol3, bcol4 = st.columns(4)
with bcol1:
    bank_files["MCB"] = st.file_uploader("MCB", type=["xlsx"], key="mcb_upload")
with bcol2:
    bank_files["HBL"] = st.file_uploader("HBL", type=["xlsx"], key="hbl_upload")
with bcol3:
    bank_files["UBL"] = st.file_uploader("UBL", type=["xlsx"], key="ubl_upload")
with bcol4:
    bank_files["FAYSAL"] = st.file_uploader("Faysal", type=["xlsx"], key="faysal_upload")

st.divider()

if st.button("🔄 Run Reconciliation"):

    if gl_file is None:
        st.warning("Please upload the GL file first.")
    else:
        try:
            gl_df = load_gl(gl_file)
        except Exception as e:
            st.error(f"Could not read GL file: {e}")
            st.stop()

        results = []
        errors = []

        for bank_key, file in bank_files.items():
            if file is None:
                continue
            try:
                bank_df = load_bank(file, bank_key)
                result = reconcile_account(gl_df, bank_df, bank_key, depot_code)
                results.append(result)
            except Exception as e:
                errors.append(f"{bank_key}: {e}")

        for err in errors:
            st.error(f"Error processing {err}")

        if not results:
            st.warning("No bank files were successfully processed. Upload at least one bank file.")
        else:
            st.session_state["recon_results"] = results
            st.session_state["recon_depot"] = depot_code
            st.success(f"Reconciliation complete for depot {depot_code}")

if "recon_results" in st.session_state:

    results = st.session_state["recon_results"]
    depot_code = st.session_state["recon_depot"]

    st.subheader("Summary")

    summary_data = []
    for r in results:
        total = len(r["ledger_debits"])
        matched = (r["ledger_debits"]["Status"] == "MATCHED").sum()
        summary_data.append({
            "Bank": r["account_label"],
            "Ledger Items": total,
            "Matched": matched,
            "Needs Review": total - matched,
            "Unmatched Bank Items": len(r["unmatched_bank_items"]),
        })
    st.dataframe(pd.DataFrame(summary_data), use_container_width=True, hide_index=True)

    st.subheader("Details by Bank")

    for r in results:
        with st.expander(f"{r['account_label']} — {len(r['ledger_debits'])} ledger items"):

            dr_display = r["ledger_debits"][
                [c for c in ["Journal Line Number", "Journal Line Description", "Entered Amount DR",
                             "Status", "Bank Deposit Slip", "Bank Amount", "Bank Date"]
                 if c in r["ledger_debits"].columns]
            ]

            def highlight_status(row):
                if row["Status"] == "UNMATCHED":
                    return ["background-color: #fdeaea"] * len(row)
                return ["background-color: #e6f4ea"] * len(row)

            st.dataframe(
                dr_display.style.apply(highlight_status, axis=1),
                use_container_width=True, hide_index=True
            )

            if len(r["unmatched_bank_items"]) > 0:
                st.markdown("**Bank items for this depot with no matching ledger entry:**")
                st.dataframe(r["unmatched_bank_items"], use_container_width=True, hide_index=True)

    st.divider()

    results_by_bank = {r["bank"]: r for r in results}
    template_excel = build_template_excel(results_by_bank, depot_code)

    st.download_button(
        "⬇️ Download Reconciliation (Main Recon + Annex 1 + Annex 3 format)",
        data=template_excel,
        file_name=f"BR_Depot_{depot_code}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    simple_excel = build_excel(results, depot_code)

    st.download_button(
        "⬇️ Download Simple Summary (one sheet per bank, easier to skim)",
        data=simple_excel,
        file_name=f"Bank_Reconciliation_Depot_{depot_code}_Simple.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
